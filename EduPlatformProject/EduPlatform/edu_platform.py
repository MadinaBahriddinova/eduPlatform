import json
from models.users import User, Student, Teacher, Parent, Admin
from models.assignments import Assignment
from models.grades import Grade
from models.schedules import Schedule
from core.enums import UserRole, AssignmentDifficulty, AssignmentStatus
import csv
import openpyxl
import datetime # For export logging
import os
class EduPlatform:
    """
    The central class for the EduPlatform.
    Manages all in-memory data (users, assignments, grades, schedules)
    and provides methods for system-wide operations.
    """
    def __init__(self):
        # Initialize data stores
        self.users = {}       # Stores User objects by ID
        self.assignments = {} # Stores Assignment objects by ID
        self.grades = {}      # Stores Grade objects by ID
        self.schedules = {}   # Stores Schedule objects by ID
        
        # Initialize counters for next available IDs (if not handled by individual models)
        # It's better if models manage their own _next_id, but if EduPlatform assigns them,
        # you'd have self.next_user_id, self.next_assignment_id, etc.
        # Assuming your models already have _next_id for now.

        self.export_log = [] # Initialize a list to store export events

        # Define base export directories relative to where the script is run
        self.EXPORT_CSV_DIR = "exported_csv"
        self.EXPORT_XLSX_DIR = "exported_xlsx"
        self.EXPORT_SQL_DIR = "exported_sql"

        # Create export directories if they don't exist
        os.makedirs(self.EXPORT_CSV_DIR, exist_ok=True)
        os.makedirs(self.EXPORT_XLSX_DIR, exist_ok=True)
        os.makedirs(self.EXPORT_SQL_DIR, exist_ok=True)

        # Initialize the admin user at startup
        self._initialize_admin()

    def _initialize_admin(self):
        """Initializes a default admin user if no users exist."""
        if not self.users:
            admin_user = Admin("Super Admin", "admin@eduplatform.com", "adminpass")
            self.users[admin_user.id] = admin_user
            print(f"Initial Admin user created: {admin_user.full_name} (ID: {admin_user.id})")

    # --- User Management Methods ---
    def register_user(self, full_name: str, email: str, password: str, role: UserRole, **kwargs) -> User | None:
        """
        Registers a new user into the platform.
        """
        # Check for existing email
        for user in self.users.values():
            if user.email == email:
                print(f"Error: User with email '{email}' already exists.")
                return None

        new_user = None
        if role == UserRole.STUDENT:
            grade_level = kwargs.get("grade_level")
            if not grade_level:
                print("Error: Student registration requires 'grade_level'.")
                return None
            new_user = Student(full_name, email, password, grade_level)
        elif role == UserRole.TEACHER:
            new_user = Teacher(full_name, email, password)
        elif role == UserRole.PARENT:
            new_user = Parent(full_name, email, password)
            children_ids = kwargs.get("children_ids")
            if children_ids and isinstance(children_ids, list):
                for child_id in children_ids:
                    if child_id in self.users and isinstance(self.users[child_id], Student):
                        new_user.children.append(child_id)
                    else:
                        print(f"Warning: Child ID {child_id} not found or not a student.")
        elif role == UserRole.ADMIN:
            new_user = Admin(full_name, email, password)
        else:
            print(f"Error: Invalid user role '{role.value}'.")
            return None

        if new_user:
            self.users[new_user.id] = new_user
            print(f"Successfully registered {role.value}: {full_name} (ID: {new_user.id})")
            return new_user
        return None

    def get_user(self, user_id: int) -> User | None:
        """Retrieves a user by their ID."""
        return self.users.get(user_id)

    def authenticate_user(self, email: str, password: str) -> User | None:
        """Authenticates a user and returns the user object if successful."""
        for user in self.users.values():
            if user.email == email:
                if user.authenticate(password):
                    print(f"Authentication successful for {user.full_name} ({user.role.value}).")
                    return user
                else:
                    print("Authentication failed: Incorrect password.")
                    return None
        print(f"Authentication failed: User with email '{email}' not found.")
        return None

    def remove_user(self, user_id: int) -> bool:
        """Removes a user from the platform and cleans up related data."""
        if user_id not in self.users:
            print(f"Error: User with ID {user_id} not found.")
            return False

        user_to_remove = self.users[user_id]
        user_role = user_to_remove.role

        # Clean up related data based on role (simple cleanup for now)
        if user_role == UserRole.STUDENT:
            # Remove student's grades and submissions from assignments
            for assignment in self.assignments.values():
                if user_id in assignment.submissions:
                    del assignment.submissions[user_id]
                if user_id in assignment.grades:
                    del assignment.grades[user_id]
            # Remove grades directly associated with this student
            self.grades = {g_id: g for g_id, g in self.grades.items() if g.student_id != user_id}
            # Remove student from parent's children list
            for parent_user in self.users.values():
                if isinstance(parent_user, Parent) and user_id in parent_user.children:
                    parent_user.children.remove(user_id)

        elif user_role == UserRole.TEACHER:
            # Remove assignments created by this teacher
            assignments_to_remove_ids = [aid for aid, a in self.assignments.items() if a.teacher_id == user_id]
            for aid in assignments_to_remove_ids:
                del self.assignments[aid]
            # Remove grades given by this teacher
            self.grades = {g_id: g for g_id, g in self.grades.items() if g.teacher_id != user_id}
            # Consider removing teacher from schedules, though more complex.

        elif user_role == UserRole.PARENT:
            # No specific data to remove directly linked to parent beyond their own object
            pass # Children references are managed from the Student side

        del self.users[user_id]
        print(f"User {user_to_remove.full_name} (ID: {user_id}, Role: {user_role.value}) removed successfully.")
        return True

    # --- Assignment Management Methods ---
    def create_assignment(self, teacher_id: int, title: str, description: str, deadline: str,
                          subject: str, class_id: str, difficulty: AssignmentDifficulty) -> Assignment | None:
        """
        Creates an assignment through a teacher.
        """
        teacher = self.get_user(teacher_id)
        if not isinstance(teacher, Teacher):
            print(f"Error: User ID {teacher_id} is not a teacher.")
            return None
        
        # Teacher's create_assignment method now just validates and creates the object
        new_assignment = teacher.create_assignment(
            title, description, deadline, subject, class_id, difficulty, self.assignments # Pass self.assignments
        )
        # The assignment is already added to self.assignments and teacher.assignments by the teacher method
        
        # Notify relevant students and parents
        for user_id, user_obj in self.users.items():
            if isinstance(user_obj, Student) and user_obj.grade == class_id:
                user_obj.add_notification(
                    f"New assignment: '{title}' in {subject} due by {deadline}.", priority="important"
                )
                # Also notify their parents
                for parent_user in self.users.values():
                    if isinstance(parent_user, Parent) and user_obj.id in parent_user.children and \
                       parent_user.notification_preferences.get("new_assignment_alert", True):
                       parent_user.add_notification(
                           f"New assignment for {user_obj.full_name}: '{title}' in {subject}.", priority="normal"
                       )
        self.export_data_on_change() # Automatic export
        return new_assignment

    def submit_assignment(self, student_id: int, assignment_id: int, content: str) -> bool:
        """
        Student submits an assignment.
        """
        student = self.get_user(student_id)
        if not isinstance(student, Student):
            print(f"Error: User ID {student_id} is not a student.")
            return False

        assignment = self.assignments.get(assignment_id)
        if not assignment:
            print(f"Error: Assignment ID {assignment_id} not found.")
            return False
        
        result = student.submit_assignment(assignment_id, content, self.assignments) # Pass self.assignments
        self.export_data_on_change() # Automatic export
        return result

    # --- Grading Management Methods ---
    def grade_assignment(self, teacher_id: int, student_id: int, assignment_id: int, grade_value: int, comment: str = None) -> bool:
        """
        Teacher grades a student's assignment.
        """
        teacher = self.get_user(teacher_id)
        if not isinstance(teacher, Teacher):
            print(f"Error: User ID {teacher_id} is not a teacher.")
            return False

        student = self.get_user(student_id)
        if not isinstance(student, Student):
            print(f"Error: User ID {student_id} is not a student.")
            return False

        assignment = self.assignments.get(assignment_id)
        if not assignment:
            print(f"Error: Assignment ID {assignment_id} not found.")
            return False

        # Ensure the teacher is authorized to grade this subject/assignment if needed
        # (e.g., check if teacher.subjects contains assignment.subject)
        if assignment.subject not in teacher.subjects:
             print(f"Error: Teacher {teacher.full_name} does not teach {assignment.subject}.")
             return False

        # Use the Teacher's method to handle the grading logic
        success = teacher.grade_assignment(assignment_id, student_id, grade_value, self.assignments, self.users)
        
        if success:
            # Create a Grade object and add it to the platform's global grades
            new_grade_obj = Grade(student_id, assignment.subject, grade_value, teacher_id, comment)
            self.grades[new_grade_obj.id] = new_grade_obj

            # Notify student about the grade
            student.add_notification(
                f"You received a grade of {grade_value} for assignment '{assignment.title}' in {assignment.subject}.",
                priority="important"
            )
            # Notify parent about the grade if notification preference is on and grade is low
            if grade_value <= 2: # Example: Notify for low grades (2 or less)
                for parent_user in self.users.values():
                    if isinstance(parent_user, Parent) and student_id in parent_user.children and \
                       parent_user.notification_preferences.get("low_grade_alert", True):
                       parent_user.add_notification(
                           f"Warning: Your child {student.full_name} received a low grade ({grade_value}) for '{assignment.title}'.",
                           priority="important"
                       )
        self.export_data_on_change() # Automatic export
        return success

    # --- Schedule Management Methods ---
    def create_schedule(self, class_id: str, day: str) -> Schedule | None:
        """Creates a new schedule for a class on a specific day."""
        # Check if schedule for this class and day already exists
        for sched in self.schedules.values():
            if sched.class_id == class_id and sched.day.lower() == day.lower():
                print(f"Error: Schedule for class '{class_id}' on '{day}' already exists.")
                return None

        new_schedule = Schedule(class_id, day)
        self.schedules[new_schedule.id] = new_schedule
        print(f"Schedule created for class {class_id} on {day} (ID: {new_schedule.id}).")
        self.export_data_on_change() # Automatic export
        return new_schedule

    def add_lesson_to_schedule(self, schedule_id: int, time: str, subject: str, teacher_id: int) -> bool:
        """Adds a lesson to an existing schedule."""
        schedule = self.schedules.get(schedule_id)
        if not schedule:
            print(f"Error: Schedule with ID {schedule_id} not found.")
            return False
        
        teacher = self.get_user(teacher_id)
        if not isinstance(teacher, Teacher):
            print(f"Error: Teacher with ID {teacher_id} not found or is not a teacher.")
            return False

        # Check for teacher's availability (optimization)
        for sched in self.schedules.values():
            if sched.id != schedule_id: # Don't check against the same schedule being modified
                for lesson_time, lesson_info in sched.lessons.items():
                    if lesson_info['teacher_id'] == teacher_id and lesson_time == time and sched.day == schedule.day:
                        print(f"Error: Teacher {teacher.full_name} is already scheduled for {sched.class_id} "
                              f"at {time} on {schedule.day}.")
                        return False

        success = schedule.add_lesson(time, subject, teacher_id)
        self.export_data_on_change() # Automatic export
        return success

    # --- Reporting Methods (Admin specific, but managed by EduPlatform) ---
    def generate_report(self, admin_id: int, report_type: str) -> dict | None:
        """
        Admin generates a report.
        """
        admin_user = self.get_user(admin_id)
        if not isinstance(admin_user, Admin):
            print(f"Error: User ID {admin_id} is not an Admin and cannot generate reports.")
            return None
        
        # Pass all relevant data collections to the admin's report method
        all_data = {
            "users": self.users,
            "assignments": self.assignments,
            "grades": self.grades,
            "schedules": self.schedules
        }
        return admin_user.generate_report(report_type, all_data)

    # --- Data Export Methods ---
    def _get_data_for_export(self) -> dict:
        """Helper to collect data for export."""
        return {
            "users": [user.get_profile() for user in self.users.values()],
            "assignments": [assign.get_assignment_info() for assign in self.assignments.values()],
            "grades": [grade.get_grade_info() for grade in self.grades.values()],
            "schedules": [schedule.view_schedule() for schedule in self.schedules.values()]
        }
        
    def _log_export_event(self, export_type: str, file_name: str):
        """Logs an export event with a timestamp."""
        timestamp = datetime.datetime.now().isoformat()
        self.export_log.append({
            "timestamp": timestamp,
            "type": export_type,
            "file_name": file_name
        })
        # print(f"Logged export event: Type='{export_type}', File='{file_name}'") # Optional: For debugging

    # Add this new method to your EduPlatform class
    def save_export_log(self, log_file_name: str = "export_log.txt"):
        """
        Saves the current export log to a text file.
        Each log entry is written as a JSON string on a new line.
        """
        try:
            # We can save it directly in the project root or in a 'logs' folder
            # For simplicity, let's put it in the project root for now.
            # If you want it in a subfolder, e.g., 'logs/export_log.txt',
            # you'd need to create that folder: os.makedirs("logs", exist_ok=True)
            # and then: log_file_path = os.path.join("logs", log_file_name)
            log_file_path = log_file_name # Saves in the current working directory (project root)

            with open(log_file_path, 'w', encoding='utf-8') as f:
                for entry in self.export_log:
                    f.write(json.dumps(entry) + '\n') # Write each dictionary as a JSON string

            print(f"Export log successfully saved to {log_file_path}")
        except Exception as e:
            print(f"Error saving export log: {e}")
            import traceback
            traceback.print_exc()

    def export_to_xlsx(self, filename="eduplatform_data.xlsx"):
        """Exports all data to an XLSX file, with each table on a separate sheet."""
        data_to_export = self._get_data_for_export()
        workbook = openpyxl.Workbook()

        first_sheet = True
        for sheet_name, records in data_to_export.items():
            if first_sheet:
                sheet = workbook.active
                sheet.title = sheet_name.capitalize()
                first_sheet = False
            else:
                sheet = workbook.create_sheet(title=sheet_name.capitalize())

            if not records:
                sheet.append(["No data"])
                continue

            # Write header
            headers = list(records[0].keys())
            sheet.append(headers)

            # Write data rows
            for record in records:
                row_data = [
                    # This is where the conversion happens.
                    # Ensure values are simple types or strings.
                    # The modification in Assignment.get_assignment_info handles this.
                    record.get(header) for header in headers
                ]
                sheet.append(row_data) # This line was failing because of [] or {}
        
        try:
            workbook.save(filename)
            print(f"Data successfully exported to {filename} (XLSX).")
            self.export_log.append(f"[{datetime.datetime.now().isoformat()}] Exported to XLSX: {filename}")
        except Exception as e:
            print(f"Error exporting to XLSX: {e}")

    def export_to_csv(self):
        try:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

            # --- User Export (The problematic section) ---
            users_data = [user.get_profile() for user in self.users.values()]
            if users_data:
                # Dynamically collect all unique fieldnames from all user profiles
                all_user_fieldnames = sorted(list(set(key for user_dict in users_data for key in user_dict.keys())))
                
                users_file_name = f"eduplatform_data_users_{timestamp}.csv"
                with open(users_file_name, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.DictWriter(csvfile, fieldnames=all_user_fieldnames)
                    writer.writeheader()
                    for user_dict in users_data:
                        # Ensure all fields are present, fill missing with empty string
                        row_to_write = {field: user_dict.get(field, '') for field in all_user_fieldnames}
                        writer.writerow(row_to_write)
                print(f"Data for users successfully exported to {users_file_name} (CSV).")
            else:
                users_file_name = f"eduplatform_data_users_{timestamp}.csv"
                open(users_file_name, 'a').close() # Create empty file
                print(f"No data for users, created empty {users_file_name}.")


            # ... (Rest of your CSV export logic for assignments, grades, schedules) ...

            # Example for assignments (already working, but shown for context)
            assignments_data = [assign.get_assignment_info() for assign in self.assignments.values()]
            if assignments_data:
                assignments_fieldnames = sorted(list(assignments_data[0].keys()))
                assignments_file_name = f"eduplatform_data_assignments_{timestamp}.csv"
                with open(assignments_file_name, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.DictWriter(csvfile, fieldnames=assignments_fieldnames)
                    writer.writeheader()
                    writer.writerows(assignments_data)
                print(f"Data for assignments successfully exported to {assignments_file_name} (CSV).")
            else:
                assignments_file_name = f"eduplatform_data_assignments_{timestamp}.csv"
                open(assignments_file_name, 'a').close()
                print(f"No data for assignments, created empty {assignments_file_name}.")
            
            # ... (grades and schedules logic) ...

            self._log_export_event("csv", f"users_{timestamp}.csv, assignments_{timestamp}.csv, etc.")
            print("--- Auto-export complete ---")

        except Exception as e:
            print(f"Error exporting to CSV: {e}")
            import traceback
            traceback.print_exc()

    def export_to_sql(self, filename="eduplatform_data.sql"):
        """Generates SQL CREATE TABLE and INSERT INTO statements for SSMS."""
        data_to_export = self._get_data_for_export()

        sql_statements = []

        # Mapping Python types to SQL Server types (simplified)
        type_mapping = {
            int: "INT",
            str: "NVARCHAR(255)", # NVARCHAR for string flexibility
            bool: "BIT",
            float: "FLOAT",
            dict: "NVARCHAR(MAX)", # Store dicts as JSON strings
            list: "NVARCHAR(MAX)" # Store lists as JSON strings
        }

        for table_name, records in data_to_export.items():
            sanitized_table_name = f"tbl_{table_name.rstrip('s')}" # e.g., users -> tbl_user, assignments -> tbl_assignment
            
            if not records:
                sql_statements.append(f"-- No data for {sanitized_table_name}\n")
                continue

            # Generate CREATE TABLE statement
            columns = []
            record_keys = list(records[0].keys()) # Use keys from first record for column names
            for key in record_keys:
                sample_value = records[0].get(key)
                # Handle potential None values or empty records in a robust way
                if sample_value is None and len(records) > 1:
                    # Try to find a non-None sample for type inference
                    for rec in records:
                        if rec.get(key) is not None:
                            sample_value = rec.get(key)
                            break

                col_type = "NVARCHAR(MAX)" # Default to NVARCHAR(MAX) if type cannot be inferred
                if sample_value is not None:
                    col_type = type_mapping.get(type(sample_value), "NVARCHAR(MAX)")
                
                # Add PRIMARY KEY for 'id' column
                if key == 'id':
                    columns.append(f"[{key}] {col_type} PRIMARY KEY")
                else:
                    columns.append(f"[{key}] {col_type}")
            
            create_table_sql = f"CREATE TABLE [{sanitized_table_name}] (\n    " + ",\n    ".join(columns) + "\n);\n"
            sql_statements.append(create_table_sql)

            # Generate INSERT INTO statements
            for record in records:
                # Prepare values for SQL (handle None, strings, and complex types)
                values = []
                for key in record_keys: # Ensure consistent order of values
                    value = record.get(key)
                    if value is None:
                        values.append("NULL")
                    elif isinstance(value, (int, float)):
                        values.append(str(value))
                    elif isinstance(value, (dict, list)):
                        # Convert dict/list to JSON string for NVARCHAR(MAX)
                        import json
                        values.append(f"N'{json.dumps(value)}'")
                    else: # Assume string
                        # Escape single quotes within strings
                        escaped_value = str(value).replace("'", "''")
                        values.append(f"N'{escaped_value}'")
                
                insert_into_sql = f"INSERT INTO [{sanitized_table_name}] ({', '.join([f'[{k}]' for k in record_keys])}) VALUES ({', '.join(values)});\n"
                sql_statements.append(insert_into_sql)
            
            sql_statements.append("\nGO\n\n") # Separator for SSMS batches

        try:
            with open(filename, 'w', encoding='utf-8') as f:
                f.writelines(sql_statements)
            print(f"SQL statements successfully exported to {filename}.")
            self.export_log.append(f"[{datetime.datetime.now().isoformat()}] Exported to SQL: {filename}")
        except Exception as e:
            print(f"Error exporting to SQL: {e}")

    def export_data_on_change(self):
        """
        Automatically exports data to XLSX, CSV, and SQL whenever a significant change occurs.
        This simple version exports all data every time.
        Optimization for "only new/changed data" would require tracking changes per object.
        """
        print("\n--- Auto-exporting data due to change ---")
        self.export_to_xlsx()
        self.export_to_csv()
        self.export_to_sql()
        print("--- Auto-export complete ---\n")

    def view_export_log(self):
        """Displays the log of export operations."""
        print("\n--- Export Log ---")
        if not self.export_log:
            print("No export operations logged yet.")
        for log_entry in self.export_log:
            print(log_entry)
        print("------------------")

    def _log_export_event(self, export_type: str, file_name: str):
        """Logs an export event with a timestamp."""
        timestamp = datetime.datetime.now().isoformat()
        self.export_log.append({
            "timestamp": timestamp,
            "type": export_type,
            "file_name": file_name
        })
        # print(f"Logged export event: Type='{export_type}', File='{file_name}'") # For debugging

    # ... (rest of your export_to_csv, export_to_xlsx, export_to_sql methods) ...

    # Example call from export_to_csv, ensure this method exists:
    
    def export_to_xlsx(self):
        try:
            workbook = openpyxl.Workbook()

            # ... (your existing XLSX export logic for sheets) ...

            # Change this line:
            # output_file = "eduplatform_data.xlsx"
            # To this:
            output_file = os.path.join(self.EXPORT_XLSX_DIR, "eduplatform_data.xlsx")

            workbook.save(output_file)
            print(f"Data successfully exported to {output_file} (XLSX).")
            self._log_export_event("xlsx", output_file) # Log the path used

        except Exception as e:
            print(f"Error exporting to XLSX: {e}")
            import traceback
            traceback.print_exc()


    def export_to_csv(self):
        try:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

            # --- Users CSV Export ---
            users_data = [user.get_profile() for user in self.users.values()]
            # Change this line:
            # users_file_name = f"eduplatform_data_users_{timestamp}.csv"
            # To this:
            users_file_name = os.path.join(self.EXPORT_CSV_DIR, f"eduplatform_data_users_{timestamp}.csv")
            
            # ... (rest of users export logic) ...

            # --- Assignments CSV Export ---
            assignments_data = [assign.get_assignment_info() for assign in self.assignments.values()]
            # Change this line:
            # assignments_file_name = f"eduplatform_data_assignments_{timestamp}.csv"
            # To this:
            assignments_file_name = os.path.join(self.EXPORT_CSV_DIR, f"eduplatform_data_assignments_{timestamp}.csv")
            
            # ... (rest of assignments export logic) ...

            # --- Grades CSV Export ---
            grades_data = [grade.get_grade_info() for grade in self.grades.values()]
            # Change this line:
            # grades_file_name = f"eduplatform_data_grades_{timestamp}.csv"
            # To this:
            grades_file_name = os.path.join(self.EXPORT_CSV_DIR, f"eduplatform_data_grades_{timestamp}.csv")
            
            # ... (rest of grades export logic) ...

            # --- Schedules CSV Export ---
            schedules_data = [schedule.get_schedule_info() for schedule in self.schedules.values()]
            # Change this line:
            # schedules_file_name = f"eduplatform_data_schedules_{timestamp}.csv"
            # To this:
            schedules_file_name = os.path.join(self.EXPORT_CSV_DIR, f"eduplatform_data_schedules_{timestamp}.csv")
            
            # ... (rest of schedules export logic) ...

            # Adjust the _log_export_event message to reflect the new paths
            self._log_export_event("csv", f"{self.EXPORT_CSV_DIR}/users_{timestamp}.csv, {self.EXPORT_CSV_DIR}/assignments_{timestamp}.csv, {self.EXPORT_CSV_DIR}/grades_{timestamp}.csv, {self.EXPORT_CSV_DIR}/schedules_{timestamp}.csv")
            print("--- Auto-export complete ---")

        except Exception as e:
            print(f"Error exporting to CSV: {e}")
            import traceback
            traceback.print_exc()




    def export_to_sql(self):
        # Change this line:
        # output_file = "eduplatform_data.sql"
        # To this:
        output_file = os.path.join(self.EXPORT_SQL_DIR, "eduplatform_data.sql")
        
        try:
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write("-- SQL Export for EduPlatform Data\n\n")

                # ... (your existing SQL generation logic) ...

            print(f"SQL statements successfully exported to {output_file}.")
            self._log_export_event("sql", output_file) # Log the path used

        except Exception as e:
            print(f"Error exporting to SQL: {e}")
            import traceback
            traceback.print_exc()